from flask import Blueprint, request, jsonify, session, send_file, after_this_request
import os
import io
from .auth import login_required
from .companies import get_companies_data
import logging
from datetime import datetime, timezone, timedelta
from reportlab.pdfgen import canvas
from reportlab.lib import colors

exports_bp = Blueprint('exports', __name__)
logger = logging.getLogger(__name__)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

def get_local_now():
    """Retorna datetime atual ajustado para o horário oficial de Brasília (UTC-3)."""
    return datetime.now(timezone(timedelta(hours=-3)))

def add_history_event(evento, detalhes):
    from ..models import SyncHistory
    SyncHistory.log_event(evento, detalhes)

class NumberedCanvas(canvas.Canvas):
    """Canvas customizado para desenhar cabeçalhos e rodapés com paginação 'Página X de Y'."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Header (páginas após a primeira)
        if self._pageNumber > 1:
            self.drawString(30, 565, "REGISTRALE · Diário Oficial da União (DOU)")
            self.drawRightString(812, 565, get_local_now().strftime('%d/%m/%Y %H:%M'))
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.75)
            self.line(30, 558, 812, 558)
            
        # Rodapé em todas as páginas
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.75)
        self.line(30, 32, 812, 32)
        
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        self.drawString(30, 20, f"Registrale Ro-DOU · Gerado automaticamente em {get_local_now().strftime('%d/%m/%Y às %H:%M:%S')} · Documento confidencial")
        page_str = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(812, 20, page_str)
        self.restoreState()

def style_excel_workbook(target):
    """Aplica formatação visual profissional às abas da planilha Excel (arquivo ou BytesIO)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        if isinstance(target, io.BytesIO):
            target.seek(0)
            wb = openpyxl.load_workbook(target)
        else:
            wb = openpyxl.load_workbook(target)
        
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        meta_header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10, color="1E293B")
        bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        for ws in wb.worksheets:
            ws.views.sheetView[0].showGridLines = True
            is_meta = "Metadados" in ws.title
            
            # Formatar primeira linha
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = meta_header_fill if is_meta else header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[1].height = 26
                
            # Formatar linhas de dados
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 20
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = bold_font if (is_meta and col == 1) else regular_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if col == 1 and not is_meta:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajuste de largura das colunas
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 70)
                
        if isinstance(target, io.BytesIO):
            target.seek(0)
            wb.save(target)
            target.seek(0)
        else:
            wb.save(target)
    except Exception as e:
        logger.error(f"Erro ao estilizar Excel: {e}")

@exports_bp.route('/export_report')
@login_required
def export_report():
    """Exporta a base de empresas cadastradas com guia dedicada de metadados."""
    import pandas as pd
    import tempfile
    
    empresas = get_companies_data()
    
    df_data = []
    for c in empresas:
        df_data.append({
            'Razão Social': c.get('nome', 'N/A'),
            'CNPJ': c.get('cnpj', 'N/A'),
            'Origem': c.get('origem', 'Manual'),
            'Monitorado': 'Sim' if c.get('status') else 'Não'
        })
        
    df_empresas = pd.DataFrame(df_data)
    if df_empresas.empty:
        df_empresas = pd.DataFrame(columns=['Razão Social', 'CNPJ', 'Origem', 'Monitorado'])
        
    meta_records = [
        {'Propriedade': 'Sistema Emissor', 'Valor': 'Registrale - Ro-DOU Dashboard'},
        {'Propriedade': 'Tipo de Relatório', 'Valor': 'Empresas Monitoradas no DOU'},
        {'Propriedade': 'Data de Geração', 'Valor': get_local_now().strftime('%d/%m/%Y')},
        {'Propriedade': 'Hora de Geração', 'Valor': get_local_now().strftime('%H:%M:%S')},
        {'Propriedade': 'Usuário Solicitante', 'Valor': f"{session.get('user', {}).get('username', 'admin')} ({session.get('user', {}).get('role', 'user')})"},
        {'Propriedade': 'Total de Empresas', 'Valor': str(len(df_empresas))},
        {'Propriedade': 'Empresas Monitoradas Ativas', 'Valor': str(len([c for c in empresas if c.get('status')]))}
    ]
    df_meta = pd.DataFrame(meta_records)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_empresas.to_excel(writer, sheet_name='Empresas Monitoradas', index=False)
        df_meta.to_excel(writer, sheet_name='Metadados da Geração', index=False)
        
    style_excel_workbook(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio_empresas.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@exports_bp.route('/test_smtp', methods=['POST'])
@login_required
def test_smtp():
    if session['user']['role'] != 'master': return jsonify({"status": "error", "message": "Acesso negado."}), 403
    data = request.json or {}
    smtp = data.get('smtp', {})
    test_email = data.get('test_email')
    
    from ..models import Settings
    settings_record = Settings.query.filter_by(key='global_settings').first()
    saved_settings = settings_record.get_value() if settings_record else {}
    saved_smtp = saved_settings.get('smtp', {}) if isinstance(saved_settings, dict) else {}
    
    server = str(smtp.get('server') or saved_smtp.get('server') or '').strip()
    port = str(smtp.get('port') or saved_smtp.get('port') or '587').strip()
    user = str(smtp.get('user') or saved_smtp.get('user') or '').strip()
    password = str(smtp.get('password') or saved_smtp.get('password') or '').strip()
    if 'gmail.com' in server.lower() or 'googlemail.com' in server.lower():
        password = password.replace(' ', '')
    from_email = str(smtp.get('from_email') or saved_smtp.get('from_email') or user).strip()
    
    if not all([server, port, test_email]):
        return jsonify({"status": "error", "message": "Servidor SMTP, porta e email de teste são obrigatórios."}), 400
        
    import smtplib
    from email.mime.text import MIMEText
    
    msg = MIMEText("Este é um email de teste enviado pelo Painel de Controle do Ro-DOU Registrale para validar as configurações de SMTP.")
    msg['Subject'] = "Ro-DOU - Teste de Conexão SMTP"
    msg['From'] = from_email
    msg['To'] = test_email
    
    server_conn = None
    try:
        port_num = int(port)
        if port_num == 465:
            server_conn = smtplib.SMTP_SSL(server, port_num, timeout=12)
        else:
            server_conn = smtplib.SMTP(server, port_num, timeout=12)
            if port_num in (587, 25):
                try:
                    server_conn.starttls()
                except Exception as tls_err:
                    logger.warning(f"STARTTLS warning: {tls_err}")
                
        if user and password:
            server_conn.login(user, password)
        server_conn.send_message(msg)
        return jsonify({"status": "success", "message": f"Email de teste enviado com sucesso para {test_email}!"})
    except smtplib.SMTPAuthenticationError as e:
        err_msg = e.smtp_error.decode('utf-8', errors='ignore') if isinstance(e.smtp_error, bytes) else str(e.smtp_error or e)
        logger.error(f"Erro de autenticação SMTP: {err_msg}")
        return jsonify({"status": "error", "message": f"Erro de autenticação SMTP: Usuário ou senha incorretos ({err_msg}). No Gmail, utilize uma 'Senha de App' de 16 dígitos sem espaços."}), 400
    except smtplib.SMTPConnectError as e:
        logger.error(f"Erro de conexão SMTP: {e}")
        return jsonify({"status": "error", "message": f"Não foi possível conectar ao servidor SMTP {server}:{port}. Verifique o endereço e a porta informados."}), 400
    except (TimeoutError, smtplib.SMTPException, OSError) as e:
        logger.error(f"Falha na comunicação SMTP: {e}")
        return jsonify({"status": "error", "message": f"Falha na conexão SMTP ({server}:{port}): {str(e)}"}), 400
    except Exception as e:
        logger.error(f"Falha inesperada ao testar SMTP: {e}")
        return jsonify({"status": "error", "message": f"Erro ao testar SMTP: {str(e)}"}), 500
    finally:
        if server_conn:
            try: server_conn.quit()
            except Exception: pass

@exports_bp.route('/send_email', methods=['POST'])
@login_required
def send_email():
    from ..services.email_service import EmailSender
    data = request.get_json(silent=True) or {}
    to_emails = data.get('to_emails', [])
    subject = data.get('subject', 'Notificação Registrale')
    body_html = data.get('body_html', '')
    
    if not to_emails:
        return jsonify({"status": "error", "message": "Nenhum destinatário informado."}), 400
        
    try:
        sender = EmailSender()
        success = sender.send_custom_email(to_emails, subject, body_html)
        if success:
            email_list = to_emails if isinstance(to_emails, list) else [str(to_emails)]
            add_history_event("Email Enviado", f"Emails enviados para: {', '.join(email_list)}")
            return jsonify({"status": "success", "message": "E-mails enviados com sucesso!"})
        else:
            return jsonify({"status": "error", "message": "Falha ao enviar e-mail via servidor SMTP."}), 500
    except Exception as e:
        logger.error(f"Erro ao enviar email: {e}")
        return jsonify({"status": "error", "message": f"Erro ao enviar email: {str(e)}"}), 500

@exports_bp.route('/export_pdf', methods=['POST'])
@login_required
def export_pdf():
    import html
    data = request.json or {}
    companies = data.get('companies', [])
    
    buffer = io.BytesIO()
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=35,
        bottomMargin=45
    )
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'HeaderSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=14
    )
    kpi_label_style = ParagraphStyle('KPILabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#64748B'), alignment=1)
    kpi_val_style = ParagraphStyle('KPIVal', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#0F172A'), alignment=1)
    
    th_style = ParagraphStyle('TH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    td_style = ParagraphStyle('TD', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, textColor=colors.HexColor('#1E293B'), leading=11)
    td_mono = ParagraphStyle('TDMono', parent=styles['Normal'], fontName='Courier', fontSize=8.5, textColor=colors.HexColor('#0F172A'), leading=11)

    # 1. Cabeçalho
    elements.append(Paragraph("Relatório de Empresas Monitoradas", title_style))
    elements.append(Paragraph("Diário Oficial da União (DOU) · Monitoramento Automatizado Registrale", subtitle_style))

    # 2. Metadados e KPIs
    active_count = len([c for c in companies if c.get('status')])
    inactive_count = len(companies) - active_count
    
    summary_data = [
        [
            Paragraph("TOTAL DE EMPRESAS", kpi_label_style),
            Paragraph("MONITORAMENTO ATIVO", kpi_label_style),
            Paragraph("MONITORAMENTO INATIVO", kpi_label_style),
            Paragraph("USUÁRIO SOLICITANTE", kpi_label_style),
            Paragraph("DATA DE GERAÇÃO", kpi_label_style)
        ],
        [
            Paragraph(str(len(companies)), kpi_val_style),
            Paragraph(f"<font color='#16a34a'>{active_count}</font>", kpi_val_style),
            Paragraph(f"<font color='#dc2626'>{inactive_count}</font>", kpi_val_style),
            Paragraph(html.escape(str(session.get('user', {}).get('username', 'admin'))), kpi_val_style),
            Paragraph(get_local_now().strftime('%d/%m/%Y %H:%M'), kpi_val_style)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[156, 156, 156, 156, 158])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))

    # 3. Tabela de Empresas
    table_data = [[
        Paragraph("Razão Social / Nome", th_style),
        Paragraph("CNPJ", th_style),
        Paragraph("Origem do Cadastro", th_style),
        Paragraph("Status", th_style)
    ]]
    
    for c in companies:
        table_data.append([
            Paragraph(html.escape(str(c.get('nome', 'N/A'))), td_style),
            Paragraph(html.escape(str(c.get('cnpj', 'N/A'))), td_mono),
            Paragraph(html.escape(str(c.get('origem', 'Manual'))), td_style),
            Paragraph("Ativo" if c.get('status') else "Inativo", td_style)
        ])
    
    if len(table_data) > 1:
        t = Table(table_data, colWidths=[360, 160, 140, 122])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Nenhuma empresa encontrada com os filtros atuais.", styles['Normal']))
        
    doc.build(elements, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="relatorio_empresas.pdf", mimetype='application/pdf')


@exports_bp.route('/export_mentions_pdf', methods=['POST'])
@login_required
def export_mentions_pdf():
    import re
    import html
    data = request.json or {}
    mentions = data.get('mentions', [])
    
    buffer = io.BytesIO()
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=35,
        bottomMargin=45
    )
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'HeaderSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=14
    )
    kpi_label_style = ParagraphStyle('KPILabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#64748B'), alignment=1)
    kpi_val_style = ParagraphStyle('KPIVal', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#0F172A'), alignment=1)
    
    th_style = ParagraphStyle('TH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    td_style = ParagraphStyle('TD', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor('#1E293B'))
    td_mono = ParagraphStyle('TDMono', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor('#475569'))
    link_style = ParagraphStyle('LinkStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=11, textColor=colors.HexColor('#2563EB'))
    
    # 1. Cabeçalho
    elements.append(Paragraph("<b>REGISTRALE</b> · Relatório Oficial de Publicações no DOU", title_style))
    elements.append(Paragraph(f"Publicações e menções identificadas no Diário Oficial da União · Emitido em {get_local_now().strftime('%d/%m/%Y às %H:%M')}", subtitle_style))
    
    # 2. Resumo Executivo
    unique_cnpjs = len(set(m.get('cnpj') for m in mentions if m.get('cnpj')))
    summary_data = [
        [
            Paragraph("TOTAL DE MENÇÕES", kpi_label_style),
            Paragraph("EMPRESAS CITADAS", kpi_label_style),
            Paragraph("USUÁRIO SOLICITANTE", kpi_label_style),
            Paragraph("DATA DA EMISSÃO", kpi_label_style)
        ],
        [
            Paragraph(str(len(mentions)), kpi_val_style),
            Paragraph(str(unique_cnpjs), kpi_val_style),
            Paragraph(html.escape(str(session.get('user', {}).get('username', 'admin'))), kpi_val_style),
            Paragraph(get_local_now().strftime('%d/%m/%Y %H:%M'), kpi_val_style)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[195, 195, 195, 197])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))
    
    # 3. Tabela de Menções
    table_data = [[
        Paragraph("Data DOU", th_style),
        Paragraph("Empresa / Razão Social", th_style),
        Paragraph("CNPJ", th_style),
        Paragraph("Seção", th_style),
        Paragraph("Trecho Identificado", th_style),
        Paragraph("Link Oficial", th_style)
    ]]
    
    for m in mentions:
        raw_trecho = m.get('trecho', '') or ''
        clean_trecho = re.sub(r'<[^>]+>', ' ', raw_trecho).strip()
        escaped_trecho = html.escape(clean_trecho)
        trecho_snippet = escaped_trecho[:160] + ('...' if len(escaped_trecho) > 160 else '')
        link = m.get('link', '')
        if link and link != '#':
            link_escaped = html.escape(link, quote=True)
            link_para = Paragraph(f'<a href="{link_escaped}">Abrir DOU</a>', link_style)
        else:
            link_para = Paragraph('-', td_style)
        
        table_data.append([
            Paragraph(html.escape(str(m.get('data', 'N/A'))), td_mono),
            Paragraph(html.escape(str(m.get('empresa', 'N/A'))), td_style),
            Paragraph(html.escape(str(m.get('cnpj', 'N/A'))), td_mono),
            Paragraph(html.escape(str(m.get('secao', 'DOU'))), td_style),
            Paragraph(trecho_snippet, td_style),
            link_para
        ])
    
    if len(table_data) > 1:
        t = Table(table_data, colWidths=[70, 150, 100, 65, 330, 67])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Nenhuma menção encontrada para os filtros selecionados.", styles['Normal']))
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="relatorio_mencoes.pdf", mimetype='application/pdf')


@exports_bp.route('/export_mentions_excel', methods=['POST'])
@login_required
def export_mentions_excel():
    """Exporta o relatório de menções com guia de dados e guia dedicada de metadados de geração."""
    import pandas as pd
    import re
    
    data = request.json or {}
    mentions = data.get('mentions', [])
    filters = data.get('filters', {})
    
    # Processar dados limpos
    processed_mentions = []
    for m in mentions:
        raw_trecho = m.get('trecho', '') or ''
        clean_trecho = re.sub(r'<[^>]+>', ' ', raw_trecho).strip()
        processed_mentions.append({
            'Data': m.get('data', 'N/A'),
            'Empresa': m.get('empresa', 'N/A'),
            'CNPJ': m.get('cnpj', 'N/A'),
            'Seção': m.get('secao', 'DOU'),
            'Trecho': clean_trecho,
            'Link Oficial': m.get('link', '')
        })
        
    df_mentions = pd.DataFrame(processed_mentions)
    if df_mentions.empty:
        df_mentions = pd.DataFrame(columns=['Data', 'Empresa', 'CNPJ', 'Seção', 'Trecho', 'Link Oficial'])
        
    # Construir guia de Metadados da Geração
    meta_records = [
        {'Propriedade': 'Sistema Emissor', 'Valor': 'Registrale - Ro-DOU Dashboard'},
        {'Propriedade': 'Tipo de Relatório', 'Valor': 'Relatório Oficial de Publicações no Diário Oficial da União'},
        {'Propriedade': 'Data de Geração', 'Valor': get_local_now().strftime('%d/%m/%Y')},
        {'Propriedade': 'Hora de Geração', 'Valor': get_local_now().strftime('%H:%M:%S')},
        {'Propriedade': 'Usuário Solicitante', 'Valor': f"{session.get('user', {}).get('username', 'admin')} ({session.get('user', {}).get('role', 'user')})"},
        {'Propriedade': 'Total de Menções Exportadas', 'Valor': str(len(df_mentions))},
        {'Propriedade': 'Empresas Únicas Citadas', 'Valor': str(len(set(m.get('cnpj') for m in mentions if m.get('cnpj'))))},
        {'Propriedade': 'Filtro de Texto', 'Valor': filters.get('search') or 'Nenhum (Todos)'},
        {'Propriedade': 'Filtro de Seção', 'Valor': filters.get('section') or 'Todas as Seções'},
        {'Propriedade': 'Data Inicial do Filtro', 'Valor': filters.get('startDate') or 'Não informada'},
        {'Propriedade': 'Data Final do Filtro', 'Valor': filters.get('endDate') or 'Não informada'}
    ]
    df_meta = pd.DataFrame(meta_records)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_mentions.to_excel(writer, sheet_name='Publicações DOU', index=False)
        df_meta.to_excel(writer, sheet_name='Metadados da Geração', index=False)
        
    style_excel_workbook(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio_mencoes.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

